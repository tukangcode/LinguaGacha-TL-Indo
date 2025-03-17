import re
import rapidjson as json
from openpyxl import Workbook
from PyQt5.QtGui import QDesktopServices
from PyQt5.QtCore import Qt
from PyQt5.QtCore import QUrl
from PyQt5.QtCore import QPoint
from PyQt5.QtWidgets import QWidget
from PyQt5.QtWidgets import QFileDialog
from PyQt5.QtWidgets import QHeaderView
from PyQt5.QtWidgets import QLayout
from PyQt5.QtWidgets import QVBoxLayout
from PyQt5.QtWidgets import QTableWidgetItem
from qfluentwidgets import Action
from qfluentwidgets import RoundMenu
from qfluentwidgets import FluentIcon
from qfluentwidgets import MessageBox
from qfluentwidgets import TableWidget
from qfluentwidgets import FluentWindow
from qfluentwidgets import TransparentPushButton

from base.Base import Base
from module.Localizer.Localizer import Localizer
from module.TableHelper import TableHelper
from widget.CommandBarCard import CommandBarCard
from widget.SwitchButtonCard import SwitchButtonCard

class PreTranslationReplacementPage(QWidget, Base):

    # 表格每列对应的数据字段
    KEYS = (
        "src",
        "dst",
    )

    def __init__(self, text: str, window: FluentWindow) -> None:
        super().__init__(window)
        self.setObjectName(text.replace(" ", "-"))

        # 根据应用语言加载默认设置
        if Localizer.get_app_language() == Base.Language.ZH:
            self.default = {
                "pre_translation_replacement_enable": False,
                "pre_translation_replacement_data": [
                    {
                        "src": "\\n[1]",
                        "dst": "张三"
                    },
                ],
            }
        else:
            self.default = {
                "pre_translation_replacement_enable": False,
                "pre_translation_replacement_data": [
                    {
                        "src": "\\n[1]",
                        "dst": "Jack"
                    },
                ],
            }

        # 载入并保存默认配置
        config = self.save_config(self.load_config_from_default())

        # 设置主容器
        self.container = QVBoxLayout(self)
        self.container.setSpacing(8)
        self.container.setContentsMargins(24, 24, 24, 24) # 左、上、右、下

        # 添加控件
        self.add_widget_head(self.container, config, window)
        self.add_widget_body(self.container, config, window)
        self.add_widget_foot(self.container, config, window)

    # 头部
    def add_widget_head(self, parent: QLayout, config: dict, window: FluentWindow) -> None:

        def init(widget: SwitchButtonCard) -> None:
            widget.set_checked(config.get("pre_translation_replacement_enable"))

        def checked_changed(widget: SwitchButtonCard, checked: bool) -> None:
            config = self.load_config()
            config["pre_translation_replacement_enable"] = checked
            self.save_config(config)

        parent.addWidget(
            SwitchButtonCard(
                Localizer.get().pre_translation_replacement_page_head_title,
                Localizer.get().pre_translation_replacement_page_head_content,
                init = init,
                checked_changed = checked_changed,
            )
        )

    # 主体
    def add_widget_body(self, parent: QLayout, config: dict, window: FluentWindow) -> None:

        def item_changed(item: QTableWidgetItem) -> None:
            item.setTextAlignment(Qt.AlignCenter)

        def insert_row(table: TableWidget) -> None:
            selected_index = self.table.selectedIndexes()

            # 有效性检验
            if selected_index == None or len(selected_index) == 0:
                return

            # 插入空行
            table.insertRow(selected_index[0].row())

        def delete_row(table: TableWidget) -> None:
            selected_index = self.table.selectedIndexes()

            # 有效性检验
            if selected_index == None or len(selected_index) == 0:
                return

            # 逆序删除并去重以避免索引错误
            for row in sorted({item.row() for item in selected_index}, reverse = True):
                table.removeRow(row)

        def custom_context_menu_requested(position: QPoint) -> None:
            menu = RoundMenu("", self.table)
            menu.addAction(
                Action(
                    FluentIcon.ADD,
                    Localizer.get().table_insert_row,
                    triggered = lambda _: insert_row(self.table),
                )
            )
            menu.addSeparator()
            menu.addAction(
                Action(
                    FluentIcon.DELETE,
                    Localizer.get().table_delete_row,
                    triggered = lambda _: delete_row(self.table),
                )
            )
            menu.exec(self.table.viewport().mapToGlobal(position))

        self.table = TableWidget(self)
        parent.addWidget(self.table)

        # 设置表格属性
        self.table.setBorderRadius(4)
        self.table.setBorderVisible(True)
        self.table.setWordWrap(False)
        self.table.setColumnCount(len(PreTranslationReplacementPage.KEYS))
        self.table.resizeRowsToContents() # 设置行高度自适应内容
        self.table.resizeColumnsToContents() # 设置列宽度自适应内容
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) # 撑满宽度
        self.table.setSelectRightClickedRow(True) # 右键选中行

        # 注册事件
        self.table.itemChanged.connect(item_changed)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(custom_context_menu_requested)

        # 设置水平表头并隐藏垂直表头
        self.table.verticalHeader().setDefaultAlignment(Qt.AlignCenter)
        self.table.setHorizontalHeaderLabels(
            (
                Localizer.get().post_translation_replacement_page_table_row_01,
                Localizer.get().post_translation_replacement_page_table_row_02,
            )
        )

        # 向表格更新数据
        TableHelper.update_to_table(self.table, config.get("pre_translation_replacement_data"), PreTranslationReplacementPage.KEYS)

    # 底部
    def add_widget_foot(self, parent: QLayout, config: dict, window: FluentWindow) -> None:
        self.command_bar_card = CommandBarCard()
        parent.addWidget(self.command_bar_card)

        # 添加命令
        self.command_bar_card.set_minimum_width(512)
        self.add_command_bar_action_import(self.command_bar_card, config, window)
        self.add_command_bar_action_export(self.command_bar_card, config, window)
        self.command_bar_card.add_separator()
        self.add_command_bar_action_add(self.command_bar_card, config, window)
        self.add_command_bar_action_save(self.command_bar_card, config, window)
        self.command_bar_card.add_separator()
        self.add_command_bar_action_reset(self.command_bar_card, config, window)
        self.command_bar_card.add_stretch(1)
        self.add_command_bar_action_wiki(self.command_bar_card, config, window)

    # 导入
    def add_command_bar_action_import(self, parent: CommandBarCard, config: dict, window: FluentWindow) -> None:

        def triggered() -> None:
            # 选择文件
            path, _ = QFileDialog.getOpenFileName(None, Localizer.get().select_file, "", Localizer.get().select_file_type)
            if not isinstance(path, str) or path == "":
                return

            # 从文件加载数据
            data = TableHelper.load_from_file(path, PreTranslationReplacementPage.KEYS)

            # 读取配置文件
            config = self.load_config()
            config["pre_translation_replacement_data"].extend(data)

            # 向表格更新数据
            TableHelper.update_to_table(self.table, config["pre_translation_replacement_data"], PreTranslationReplacementPage.KEYS)

            # 从表格加载数据（去重后）
            config["pre_translation_replacement_data"] = TableHelper.load_from_table(self.table, PreTranslationReplacementPage.KEYS)

            # 保存配置文件
            config = self.save_config(config)

            # 弹出提示
            self.emit(Base.Event.APP_TOAST_SHOW, {
                "type": Base.ToastType.SUCCESS,
                "message": Localizer.get().pre_translation_replacement_page_import_toast,
            })

        parent.add_action(
            Action(FluentIcon.DOWNLOAD, Localizer.get().pre_translation_replacement_page_import, parent, triggered = triggered),
        )

    # 导出
    def add_command_bar_action_export(self, parent: CommandBarCard, config: dict, window: FluentWindow) -> None:

        def export_to_xlsx(data: list[dict[str, str]], path: str) -> None:
            work_book = Workbook()
            active_sheet = work_book.active

            # 将数据写入工作表
            for row, item in enumerate(data):
                # 如果文本是以 = 开始，则加一个空格
                # 因为 = 开头会被识别成 Excel 公式
                src: str = re.sub(r"^=", " =", item.get("src", ""))
                dst: str = re.sub(r"^=", " =", item.get("dst", ""))
                info: str = re.sub(r"^=", " =", item.get("info", ""))

                if src != "":
                    try:
                        active_sheet.cell(row = row + 1, column = 1).value = src
                    except:
                        active_sheet.cell(row = row + 1, column = 1).value = re.escape(src)

                if dst != "":
                    try:
                        active_sheet.cell(row = row + 1, column = 2).value = dst
                    except:
                        active_sheet.cell(row = row + 1, column = 2).value = re.escape(dst)

                if info != "":
                    try:
                        active_sheet.cell(row = row + 1, column = 3).value = info
                    except:
                        active_sheet.cell(row = row + 1, column = 3).value = re.escape(info)

            # 保存工作簿
            work_book.save(path)

        def triggered() -> None:
            # 从表格加载数据
            data = TableHelper.load_from_table(self.table, PreTranslationReplacementPage.KEYS)

            # 导出文件
            export_to_xlsx(data, f"{Localizer.get().path_pre_translation_replacement_export}.xlsx")
            with open(f"{Localizer.get().path_pre_translation_replacement_export}.json", "w", encoding = "utf-8") as writer:
                writer.write(json.dumps(data, indent = 4, ensure_ascii = False))

            # 弹出提示
            self.emit(Base.Event.APP_TOAST_SHOW, {
                "type": Base.ToastType.SUCCESS,
                "message": Localizer.get().pre_translation_replacement_page_export_toast,
            })

        parent.add_action(
            Action(FluentIcon.SHARE, Localizer.get().pre_translation_replacement_page_export, parent, triggered = triggered),
        )

    # 添加新行
    def add_command_bar_action_add(self, parent: CommandBarCard, config: dict, window: FluentWindow) -> None:

        def triggered() -> None:
            # 添加新行
            self.table.setRowCount(self.table.rowCount() + 1)

            # 弹出提示
            self.emit(Base.Event.APP_TOAST_SHOW, {
                "type": Base.ToastType.SUCCESS,
                "message": Localizer.get().pre_translation_replacement_page_add_toast,
            })

        parent.add_action(
            Action(FluentIcon.ADD_TO, Localizer.get().pre_translation_replacement_page_add, parent, triggered = triggered),
        )

    # 保存
    def add_command_bar_action_save(self, parent: CommandBarCard, config: dict, window: FluentWindow) -> None:

        def triggered() -> None:
            # 加载配置文件
            config = self.load_config()

            # 从表格加载数据
            config["pre_translation_replacement_data"] = TableHelper.load_from_table(self.table, PreTranslationReplacementPage.KEYS)

            # 清空表格
            self.table.clearContents()

            # 向表格更新数据
            TableHelper.update_to_table(self.table, config["pre_translation_replacement_data"], PreTranslationReplacementPage.KEYS)

            # 从表格加载数据（去重后）
            config["pre_translation_replacement_data"] = TableHelper.load_from_table(self.table, PreTranslationReplacementPage.KEYS)

            # 保存配置文件
            config = self.save_config(config)

            # 弹出提示
            self.emit(Base.Event.APP_TOAST_SHOW, {
                "type": Base.ToastType.SUCCESS,
                "message": Localizer.get().pre_translation_replacement_page_save_toast,
            })

        parent.add_action(
            Action(FluentIcon.SAVE, Localizer.get().pre_translation_replacement_page_save, parent, triggered = triggered),
        )

    # 重置
    def add_command_bar_action_reset(self, parent: CommandBarCard, config: dict, window: FluentWindow) -> None:

        def triggered() -> None:
            message_box = MessageBox(Localizer.get().alert, Localizer.get().pre_translation_replacement_page_reset_alert, window)
            message_box.yesButton.setText(Localizer.get().confirm)
            message_box.cancelButton.setText(Localizer.get().cancel)

            if not message_box.exec():
                return

            # 清空表格
            self.table.clearContents()

            # 加载配置文件
            config = self.load_config()

            # 加载默认设置
            config["pre_translation_replacement_data"] = self.default.get("pre_translation_replacement_data")

            # 保存配置文件
            config = self.save_config(config)

            # 向表格更新数据
            TableHelper.update_to_table(self.table, config.get("pre_translation_replacement_data"), PreTranslationReplacementPage.KEYS)

            # 弹出提示
            self.emit(Base.Event.APP_TOAST_SHOW, {
                "type": Base.ToastType.SUCCESS,
                "message": Localizer.get().pre_translation_replacement_page_reset_toast,
            })

        parent.add_action(
            Action(FluentIcon.DELETE, Localizer.get().pre_translation_replacement_page_reset, parent, triggered = triggered),
        )

    # WiKi
    def add_command_bar_action_wiki(self, parent: CommandBarCard, config: dict, window: FluentWindow) -> None:

        def connect() -> None:
            QDesktopServices.openUrl(QUrl("https://github.com/neavo/LinguaGacha/wiki"))

        push_button = TransparentPushButton(FluentIcon.HELP, Localizer.get().wiki)
        push_button.clicked.connect(connect)
        parent.add_widget(push_button)