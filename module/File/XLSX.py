import os
import re

import openpyxl

from base.Base import Base
from module.Cache.CacheItem import CacheItem

class XLSX(Base):

    def __init__(self, config: dict) -> None:
        super().__init__()

        # 初始化
        self.config: dict = config
        self.input_path: str = config.get("input_folder")
        self.output_path: str = config.get("output_folder")
        self.source_language: str = config.get("source_language")
        self.target_language: str = config.get("target_language")

    # 读取
    def read_from_path(self, abs_paths: list[str]) -> list[CacheItem]:
        items = []
        for abs_path in set(abs_paths):
            # 获取相对路径
            rel_path = os.path.relpath(abs_path, self.input_path)

            # 数据处理
            wb = openpyxl.load_workbook(abs_path)
            sheet = wb.active

            # 跳过空表格
            if sheet.max_row == 0 or sheet.max_column == 0:
                continue

            for row in range(1, sheet.max_row + 1):
                src = sheet.cell(row = row, column = 1).value
                dst = sheet.cell(row = row, column = 2).value

                # 跳过读取失败的行
                # 数据不存在时为 None，存在时可能是 str int float 等多种类型
                if src is None:
                    continue

                src = str(src)
                dst = str(dst) if dst is not None else ""
                if src == "":
                    items.append(
                        CacheItem({
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": CacheItem.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.TranslationStatus.EXCLUDED,
                        })
                    )
                elif dst != "" and src != dst:
                    items.append(
                        CacheItem({
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": CacheItem.FileType.XLSX,
                            "file_path": rel_path,
                            "status": Base.TranslationStatus.EXCLUDED,
                        })
                    )
                else:
                    items.append(
                        CacheItem({
                            "src": src,
                            "dst": dst,
                            "row": row,
                            "file_type": CacheItem.FileType.XLSX,
                            "file_path": rel_path,
                            "text_type": Base.TranslationStatus.UNTRANSLATED,
                        })
                    )

        return items

    # 写入
    def write_to_path(self, items: list[CacheItem]) -> None:
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.XLSX
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for rel_path, items in data.items():
            # 新建工作表
            work_book = openpyxl.Workbook()
            active_sheet = work_book.active

            # 将数据写入工作表
            for item in items:
                src: str = re.sub(r"^=", " =", item.get_src())
                dst: str = re.sub(r"^=", " =", item.get_dst())
                row: int = item.get_row()

                # 如果文本是以 = 开始，则加一个空格
                # 因为 = 开头会被识别成 Excel 公式导致 T++ 导入时 卡住
                # 加入空格后，虽然还是不能直接导入 T++ ，但是可以手动复制粘贴
                try:
                    active_sheet.cell(row = row, column = 1).value = src
                except:
                    active_sheet.cell(row = row, column = 1).value = re.escape(src)
                try:
                    active_sheet.cell(row = row, column = 2).value = dst
                except:
                    active_sheet.cell(row = row, column = 2).value = re.escape(dst)

            # 保存工作簿
            abs_path = os.path.join(self.output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            work_book.save(abs_path)