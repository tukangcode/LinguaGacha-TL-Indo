import os
import re
import random
import shutil
import warnings
from datetime import datetime

import ebooklib
import openpyxl
import rapidjson as json
from bs4 import BeautifulSoup
from ebooklib import epub
from openpyxl import Workbook
from openpyxl.utils import escape

from base.Base import Base
from module.Cache.CacheItem import CacheItem
from module.Cache.CacheProject import CacheProject
from module.TextHelper import TextHelper

class FileHelper(Base):

    # https://github.com/aerkalov/ebooklib/issues/296
    warnings.filterwarnings(
        "ignore",
        message = "In the future version we will turn default option ignore_ncx to True."
    )
    warnings.filterwarnings(
        "ignore",
        message = "This search incorrectly ignores the root element, and will be fixed in a future version"
    )

    def __init__(self, input_path: str, output_path: str) -> None:
        super().__init__()

        # 初始化
        self.input_path = input_path
        self.output_path = output_path

    # 读
    def read_from_path(self) -> tuple[CacheProject, list[CacheItem]]:
        project = CacheProject({
            "id": f"{datetime.now().strftime("%Y%m%d_%H%M%S")}_{random.randint(100000, 999999)}",
        })

        items = []
        try:
            items.extend(self.read_from_path_txt(self.input_path, self.output_path))
            items.extend(self.read_from_path_ass(self.input_path, self.output_path))
            items.extend(self.read_from_path_srt(self.input_path, self.output_path))
            items.extend(self.read_from_path_tpp(self.input_path, self.output_path))
            items.extend(self.read_from_path_epub(self.input_path, self.output_path))
            items.extend(self.read_from_path_renpy(self.input_path, self.output_path))
            items.extend(self.read_from_path_kvjson(self.input_path, self.output_path))
            items.extend(self.read_from_path_messagejson(self.input_path, self.output_path))
        except Exception as e:
            self.error(f"文件读取失败 ... {e}", e if self.is_debug() else None)

        return project, items

    # 写
    def write_to_path(self, items: list[CacheItem]) -> None:
        try:
            self.write_to_path_txt(self.input_path, self.output_path, items)
            self.write_to_path_ass(self.input_path, self.output_path, items)
            self.write_to_path_srt(self.input_path, self.output_path, items)
            self.write_to_path_tpp(self.input_path, self.output_path, items)
            self.write_to_path_epub(self.input_path, self.output_path, items)
            self.write_to_path_renpy(self.input_path, self.output_path, items)
            self.write_to_path_kvjson(self.input_path, self.output_path, items)
            self.write_to_path_messagejson(self.input_path, self.output_path, items)
        except Exception as e:
            self.error(f"文件写入失败 ... {e}", e if self.is_debug() else None)

    # TXT
    def read_from_path_txt(self, input_path: str, output_path: str) -> list[CacheItem]:
        items = []
        for root, _, files in os.walk(input_path):
            target = [
                os.path.join(root, file).replace("\\", "/") for file in files
                if file.lower().endswith(".txt")
            ]

            for abs_path in target:
                # 获取相对路径
                rel_path = os.path.relpath(abs_path, input_path)

                # 数据处理
                with open(abs_path, "r", encoding = "utf-8") as reader:
                    for line in [line.removesuffix("\n") for line in reader.readlines()]:
                        items.append(
                            CacheItem({
                                "src": line,
                                "dst": line,
                                "row": len(items),
                                "file_type": CacheItem.FileType.TXT,
                                "file_path": rel_path,
                            })
                        )

        return items

    # TXT
    def write_to_path_txt(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.TXT
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write("\n".join([item.get_dst() for item in items]))

        # 分别处理每个文件（双语）
        for rel_path, items in data.items():
            ext = os.path.splitext(rel_path)[-1]
            abs_path = os.path.join(output_path, rel_path.replace(ext, "_双语.txt"))
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write("\n".join([
                    item.get_src() if item.get_dst() == item.get_src() else f"{item.get_src()}\n{item.get_dst()}"
                    for item in items
                ]))

    # ASS
    def read_from_path_ass(self, input_path: str, output_path: str) -> list[CacheItem]:
        # [Script Info]
        # ; This is an Advanced Sub Station Alpha v4+ script.
        # Title:
        # ScriptType: v4.00+
        # PlayDepth: 0
        # ScaledBorderAndShadow: Yes

        # [V4+ Styles]
        # Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding
        # Style: Default,Arial,20,&H00FFFFFF,&H0000FFFF,&H00000000,&H00000000,0,0,0,0,100,100,0,0,1,1,1,2,10,10,10,1

        # [Events]
        # Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text
        # Dialogue: 0,0:00:08.12,0:00:10.46,Default,,0,0,0,,にゃにゃにゃ
        # Dialogue: 0,0:00:14.00,0:00:15.88,Default,,0,0,0,,えーこの部屋一人で使\Nえるとか最高じゃん
        # Dialogue: 0,0:00:15.88,0:00:17.30,Default,,0,0,0,,えるとか最高じゃん

        items = []
        for root, _, files in os.walk(input_path):
            target = [
                os.path.join(root, file).replace("\\", "/") for file in files
                if file.lower().endswith(".ass")
            ]

            for abs_path in target:
                # 获取相对路径
                rel_path = os.path.relpath(abs_path, input_path)

                # 数据处理
                with open(abs_path, "r", encoding = "utf-8") as reader:
                    lines = [line.strip() for line in reader.readlines()]

                    # 格式字段的数量
                    in_event = False
                    format_field_num = -1
                    for line in lines:
                        # 判断是否进入事件块
                        if line == "[Events]":
                            in_event = True
                        # 在事件块中寻找格式字段
                        if in_event == True and line.startswith("Format:"):
                            format_field_num = len(line.split(",")) - 1
                            break

                    for line in lines:
                        content = ",".join(line.split(",")[format_field_num:]) if line.startswith("Dialogue:") else ""
                        extra_field = line.replace(f"{content}", "{CONTENT}") if content != "" else line

                        # 添加数据
                        items.append(
                            CacheItem({
                                "src": content.replace("\\N", "\n"),
                                "dst": content.replace("\\N", "\n"),
                                "extra_field": extra_field,
                                "row": len(items),
                                "file_type": CacheItem.FileType.ASS,
                                "file_path": rel_path,
                            })
                        )

        return items

    # ASS
    def write_to_path_ass(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # [Script Info]
        # ; This is an Advanced Sub Station Alpha v4+ script.
        # Title:
        # ScriptType: v4.00+
        # PlayDepth: 0
        # ScaledBorderAndShadow: Yes

        # [V4+ Styles]
        # Format: Name, Fontname, Fontsize, PrimaryColour, SecondaryColour, OutlineColour, BackColour, Bold, Italic, Underline, StrikeOut, ScaleX, ScaleY, Spacing, Angle, BorderStyle, Outline, Shadow, Alignment, MarginL, MarginR, MarginV, Encoding
        # Style: Default,Arial,20,&H00FFFFFF,&H0000FFFF,&H00000000,&H00000000,0,0,0,0,100,100,0,0,1,1,1,2,10,10,10,1

        # [Events]
        # Format: Layer, Start, End, Style, Name, MarginL, MarginR, MarginV, Effect, Text
        # Dialogue: 0,0:00:08.12,0:00:10.46,Default,,0,0,0,,にゃにゃにゃ
        # Dialogue: 0,0:00:14.00,0:00:15.88,Default,,0,0,0,,えーこの部屋一人で使\Nえるとか最高じゃん
        # Dialogue: 0,0:00:15.88,0:00:17.30,Default,,0,0,0,,えるとか最高じゃん

        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.ASS
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                result.append(item.get_extra_field().replace("{CONTENT}", item.get_dst().replace("\n", "\\N")))

            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write("\n".join(result))

        # 分别处理每个文件（双语）
        for rel_path, items in data.items():
            ext = os.path.splitext(rel_path)[-1]
            abs_path = os.path.join(output_path, rel_path.replace(ext, "_双语.ass"))
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                result.append(
                    item.get_extra_field().replace("{CONTENT}", "{CONTENT}\\N{CONTENT}")
                                          .replace("{CONTENT}", item.get_src().replace("\n", "\\N"), 1)
                                          .replace("{CONTENT}", item.get_dst().replace("\n", "\\N"), 1)
                )

            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write("\n".join(result))

    # SRT
    def read_from_path_srt(self, input_path: str, output_path: str) -> list[CacheItem]:
        # 1
        # 00:00:08,120 --> 00:00:10,460
        # にゃにゃにゃ

        # 2
        # 00:00:14,000 --> 00:00:15,880
        # えーこの部屋一人で使

        # 3
        # 00:00:15,880 --> 00:00:17,300
        # えるとか最高じゃん

        items = []
        for root, _, files in os.walk(input_path):
            target = [
                os.path.join(root, file).replace("\\", "/") for file in files
                if file.lower().endswith(".srt")
            ]

            for abs_path in target:
                # 获取相对路径
                rel_path = os.path.relpath(abs_path, input_path)

                # 数据处理
                with open(abs_path, "r", encoding = "utf-8") as reader:
                    chunks = re.split(r"\n{2,}", reader.read().strip())
                    for chunk in chunks:
                        lines = chunk.splitlines()

                        # 格式校验
                        # isdigit
                        # 仅返回 True 如果字符串中的所有字符都是 Unicode 数字字符（例如：0-9），不包括带有上标的数字（如 ²）或带分隔符的数字（如罗马数字）。
                        # isnumeric
                        # 返回 True 如果字符串中的所有字符都是 Unicode 数值字符，包括 Unicode 数字、带有上标的数字和其他形式的数值字符（如罗马数字）。
                        if len(lines) < 3 or not lines[0].isdigit():
                            continue

                        # 添加数据
                        if lines[-1] != "":
                            items.append(
                                CacheItem({
                                    "src": "\n".join(lines[2:]),            # 如有多行文本则用换行符拼接
                                    "dst": "\n".join(lines[2:]),            # 如有多行文本则用换行符拼接
                                    "extra_field": lines[1],
                                    "row": str(lines[0]),
                                    "file_type": CacheItem.FileType.SRT,
                                    "file_path": rel_path,
                                })
                            )

        return items

    # SRT
    def write_to_path_srt(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # 1
        # 00:00:08,120 --> 00:00:10,460
        # にゃにゃにゃ

        # 2
        # 00:00:14,000 --> 00:00:15,880
        # えーこの部屋一人で使

        # 3
        # 00:00:15,880 --> 00:00:17,300
        # えるとか最高じゃん

        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.SRT
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                result.append([
                    item.get_row(),
                    item.get_extra_field(),
                    item.get_dst(),
                ])

            with open(abs_path, "w", encoding = "utf-8") as writer:
                for item in result:
                    writer.write("\n".join(item))
                    writer.write("\n\n")

        # 分别处理每个文件（双语）
        for rel_path, items in data.items():
            ext = os.path.splitext(rel_path)[-1]
            abs_path = os.path.join(output_path, rel_path.replace(ext, "_双语.srt"))
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                result.append([
                    item.get_row(),
                    item.get_extra_field(),
                    f"{item.get_src()}\n{item.get_dst()}",
                ])

            with open(abs_path, "w", encoding = "utf-8") as writer:
                for item in result:
                    writer.write("\n".join(item))
                    writer.write("\n\n")

    # T++
    def read_from_path_tpp(self, input_path: str, output_path: str) -> list[CacheItem]:
        items = []
        for root, _, files in os.walk(input_path):
            target = [
                os.path.join(root, file).replace("\\", "/") for file in files
                if file.lower().endswith(".xlsx")
            ]

            for abs_path in target:
                # 获取相对路径
                rel_path = os.path.relpath(abs_path, input_path)

                # 数据处理
                wb = openpyxl.load_workbook(abs_path)
                sheet = wb.active

                # 跳过空表格
                if sheet.max_column == 0:
                    continue

                # 跳过不包含 T++ 表头的表格
                if sheet.cell(row = 1, column = 1).value != "Original Text":
                    continue

                for row in range(2, sheet.max_row + 1):
                    src = sheet.cell(row = row, column = 1).value
                    dst = sheet.cell(row = row, column = 2).value

                    # 跳过读取失败的行
                    if src is None:
                        continue
                    else:
                        src = str(src)

                    items.append(
                        CacheItem({
                            "src": src,
                            "dst": dst if dst != None else src,
                            "row": row,
                            "file_type": CacheItem.FileType.TPP,
                            "file_path": rel_path,
                        })
                    )

        return items

    # T++
    def write_to_path_tpp(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.TPP
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for rel_path, items in data.items():
            # 新建工作表
            work_book = Workbook()
            active_sheet = work_book.active
            active_sheet.append(
                (
                    "Original Text",
                    "Initial",
                    "Machine translation",
                    "Better translation",
                    "Best translation",
                )
            )

            # 将数据写入工作表
            for item in items:
                src: str = re.sub(r"^=", " =", item.get_src())
                dst: str = re.sub(r"^=", " =", item.get_dst())
                row: int = item.get_row()

                # 如果文本是以 = 开始，则加一个空格
                # 因为 = 开头会被识别成 Excel 公式导致 T++ 导入时 卡住
                # 加入空格后，虽然还是不能直接导入 T++ ，但是可以手动复制粘贴
                try:
                    active_sheet.cell(row = row, column = 1).value = src
                except:
                    active_sheet.cell(row = row, column = 1).value = escape(src)
                try:
                    active_sheet.cell(row = row, column = 2).value = dst
                except:
                    active_sheet.cell(row = row, column = 2).value = escape(dst)

            # 保存工作簿
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            work_book.save(abs_path)

    # EPUB
    def read_from_path_epub(self, input_path: str, output_path: str) -> list[CacheItem]:
        items = []
        for root, _, files in os.walk(input_path):
            target = [
                os.path.join(root, file).replace("\\", "/") for file in files
                if file.lower().endswith(".epub")
            ]

            for abs_path in target:
                # 获取相对路径
                rel_path = os.path.relpath(abs_path, input_path)

                # 将原始文件复制一份
                os.makedirs(os.path.dirname(f"{output_path}/cache/temp/{rel_path}"), exist_ok = True)
                shutil.copy(abs_path, f"{output_path}/cache/temp/{rel_path}")

                # 数据处理
                book = epub.read_epub(abs_path)
                for doc in book.get_items_of_type(ebooklib.ITEM_DOCUMENT):
                    id = doc.get_id()
                    bs = BeautifulSoup(doc.get_content(), "html.parser")
                    for line in str(bs).splitlines():
                        content = BeautifulSoup(line, "html.parser").get_text()
                        extra_field = line.replace(f"{content}", "{CONTENT}") if content != "" else line
                        items.append(
                            CacheItem({
                                "src": content,
                                "dst": content,
                                "tag": id,
                                "row": len(items),
                                "extra_field": extra_field,
                                "file_type": CacheItem.FileType.EPUB,
                                "file_path": rel_path,
                            })
                        )

        return items

    # EPUB
    def write_to_path_epub(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:

        def add_opacity_style(html: str) -> str:
            # 找到第一个 > 的位置
            index = html.find(">")

            # 替换第一个 >
            return html[:index] + " style=\"opacity:0.4;\">" + html[index + 1:]

        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.EPUB
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            # 按行号排序
            items = sorted(items, key = lambda x: x.get_row())

            # 读取原始文件
            book = epub.read_epub(f"{output_path}/cache/temp/{rel_path}")
            for doc in book.get_items_of_type(ebooklib.ITEM_DOCUMENT):
                # 筛选出 tag 相同的元素
                lines = []
                for item in [item for item in items if item.get_tag() == doc.get_id()]:
                    lines.append(item.get_extra_field().replace("{CONTENT}", item.get_dst()))

                # 将修改后的 HTML 内容重新填充回去
                doc.set_content("\n".join(lines).encode("utf-8"))

            # 将修改后的数据写入文件
            abs_path = f"{output_path}/{rel_path}"
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            epub.write_epub(abs_path, book, {})

        # 分别处理每个文件（双语）
        for rel_path, items in data.items():
            # 按行号排序
            items = sorted(items, key = lambda x: x.get_row())

            # 读取原始文件
            book = epub.read_epub(f"{output_path}/cache/temp/{rel_path}")
            for doc in book.get_items_of_type(ebooklib.ITEM_DOCUMENT):
                # 筛选出 tag 相同的元素
                lines = []
                for item in [item for item in items if item.get_tag() == doc.get_id()]:
                    if item.get_src() != "":
                        lines.append(add_opacity_style(item.get_extra_field()).replace("{CONTENT}", item.get_src()))
                    lines.append(item.get_extra_field().replace("{CONTENT}", item.get_dst()))

                # 将修改后的 HTML 内容重新填充回去
                doc.set_content("\n".join(lines).encode("utf-8"))

            # 将修改后的数据写入文件
            ext = os.path.splitext(rel_path)[-1]
            abs_path = os.path.join(output_path, rel_path.replace(ext, "_双语.epub"))
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            epub.write_epub(abs_path, book, {})

    # RENPY
    def read_from_path_renpy(self, input_path: str, output_path: str) -> list[CacheItem]:
        # # game/script8.rpy:16878
        # translate chinese arabialogoff_e5798d9a:
        #
        #     # lo "And you...?{w=2.3}{nw}" with dissolve
        #     lo "And you...?{w=2.3}{nw}" with dissolve
        #
        # # game/script/1-home/1-Perso_Home/elice.rpy:281
        # translate schinese elice_ask_home_f01e3240_5:
        #
        #     # e ".{w=0.5}.{w=0.5}.{w=0.5}{nw}"
        #     e ".{w=0.5}.{w=0.5}.{w=0.5}{nw}"
        #
        # # game/script8.rpy:33
        # translate chinese update08_a626b58f:
        #
        #     # "*Snorts* Fucking hell, I hate this dumpster of a place." with dis06
        #     "*Snorts* Fucking hell, I hate this dumpster of a place." with dis06
        #
        # translate chinese strings:
        #
        #     # game/script8.rpy:307
        #     old "Accompany her to the inn"
        #     new "Accompany her to the inn"
        #
        #     # game/script8.rpy:2173
        #     old "{sc=3}{size=44}Jump off the ship.{/sc}"
        #     new "{sc=3}{size=44}Jump off the ship.{/sc}"

        # 查找文本中最后一对双引号包裹的文本
        def find_content(text: str) -> str:
            matches = re.findall(r'"(.*?)"', text)
            if matches:
                return matches[-1]
            else:
                return ""

        items = []
        for root, _, files in os.walk(input_path):
            target = [
                os.path.join(root, file).replace("\\", "/") for file in files
                if file.lower().endswith(".rpy")
            ]

            for abs_path in target:
                # 获取相对路径
                rel_path = os.path.relpath(abs_path, input_path)

                # 数据处理
                with open(abs_path, "r", encoding = "utf-8") as reader:
                    lines = [line.removesuffix("\n") for line in reader.readlines()]

                skip_next = False
                for line in lines:
                    if skip_next == True:
                        skip_next = False
                        continue
                    elif (line.startswith("    # ") and line.count("\"") >= 2) or line.startswith("    old "):
                        skip_next = True
                        content = find_content(line).replace("\\n", "\n")
                        extra_field = line.replace(f"{content}", "{CONTENT}") if content != "" else line
                    else:
                        content = ""
                        extra_field = line

                    # 添加数据
                    items.append(
                        CacheItem({
                            "src": content,
                            "dst": content,
                            "extra_field": extra_field,
                            "row": len(items),
                            "file_type": CacheItem.FileType.RENPY,
                            "file_path": rel_path,
                        })
                    )

        return items

    # RENPY
    def write_to_path_renpy(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # # game/script8.rpy:16878
        # translate chinese arabialogoff_e5798d9a:
        #
        #     # lo "And you...?{w=2.3}{nw}" with dissolve
        #     lo "And you...?{w=2.3}{nw}" with dissolve
        #
        # # game/script/1-home/1-Perso_Home/elice.rpy:281
        # translate schinese elice_ask_home_f01e3240_5:
        #
        #     # e ".{w=0.5}.{w=0.5}.{w=0.5}{nw}"
        #     e ".{w=0.5}.{w=0.5}.{w=0.5}{nw}"
        #
        # # game/script8.rpy:33
        # translate chinese update08_a626b58f:
        #
        #     # "*Snorts* Fucking hell, I hate this dumpster of a place." with dis06
        #     "*Snorts* Fucking hell, I hate this dumpster of a place." with dis06
        #
        # translate chinese strings:
        #
        #     # game/script8.rpy:307
        #     old "Accompany her to the inn"
        #     new "Accompany her to the inn"
        #
        #     # game/script8.rpy:2173
        #     old "{sc=3}{size=44}Jump off the ship.{/sc}"
        #     new "{sc=3}{size=44}Jump off the ship.{/sc}"

        # 筛选
        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.RENPY
        ]

        # 按文件路径分组
        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        # 分别处理每个文件
        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                extra_field = item.get_extra_field()
                if (extra_field.startswith("    # ") and extra_field.count("\"") >= 2):
                    result.append(extra_field.replace("{CONTENT}", item.get_src()).replace("\n", "\\n"))
                    result.append(extra_field.replace("    # ", "    ").replace("{CONTENT}", item.get_dst()).replace("\n", "\\n"))
                elif extra_field.startswith("    old "):
                    result.append(extra_field.replace("{CONTENT}", item.get_src()).replace("\n", "\\n"))
                    result.append(extra_field.replace("    old ", "    new ").replace("{CONTENT}", item.get_dst()).replace("\n", "\\n"))
                else:
                    result.append(extra_field)

            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write("\n".join(result))

    # KV JSON
    def read_from_path_kvjson(self, input_path: str, output_path: str) -> list[CacheItem]:
        # {
        #     "「あ・・」": "「あ・・」",
        #     "「ごめん、ここ使う？」": "「ごめん、ここ使う？」",
        #     "「じゃあ・・私は帰るね」": "「じゃあ・・私は帰るね」",
        # }

        items = []
        for root, _, files in os.walk(input_path):
            target = [
                os.path.join(root, file).replace("\\", "/") for file in files
                if file.lower().endswith(".json")
            ]

            for abs_path in target:
                # 获取相对路径
                rel_path = os.path.relpath(abs_path, input_path)

                # 数据处理
                with open(abs_path, "r", encoding = "utf-8") as reader:
                    json_data: dict[str, str] = TextHelper.safe_load_json_dict(reader.read().strip())

                    # 格式校验
                    if json_data == {}:
                        continue

                    # 读取数据
                    for k, v in json_data.items():
                        # 格式校验
                        if not isinstance(k, str) or not isinstance(v, str):
                            continue

                        if k != "":
                            items.append(
                                CacheItem({
                                    "src": k,
                                    "dst": v,
                                    "row": len(items),
                                    "file_type": CacheItem.FileType.KVJSON,
                                    "file_path": rel_path,
                                })
                            )

        return items

    # KV JSON
    def write_to_path_kvjson(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # {
        #     "「あ・・」": "「あ・・」",
        #     "「ごめん、ここ使う？」": "「ごめん、ここ使う？」",
        #     "「じゃあ・・私は帰るね」": "「じゃあ・・私は帰るね」",
        # }

        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.KVJSON
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)
            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write(
                    json.dumps(
                        {
                            item.get_src(): item.get_dst() for item in items
                        },
                        indent = 4,
                        ensure_ascii = False,
                    )
                )

    # Message JSON
    def read_from_path_messagejson(self, input_path: str, output_path: str) -> list[CacheItem]:
        # [
        #     {
        #         "message": "<fgName:pipo-fog004><fgLoopX:1><fgLoopY:1><fgSx:-2><fgSy:0.5>"
        #     },
        #     {
        #         "message": "エンディングを変更しますか？"
        #     },
        #     {
        #         "message": "はい"
        #     },
        # ]

        items = []
        for root, _, files in os.walk(input_path):
            target = [
                os.path.join(root, file).replace("\\", "/") for file in files
                if file.lower().endswith(".json")
            ]

            for abs_path in target:
                # 获取相对路径
                rel_path = os.path.relpath(abs_path, input_path)

                # 数据处理
                with open(abs_path, "r", encoding = "utf-8") as reader:
                    json_data: list[dict] = TextHelper.safe_load_json_list(reader.read().strip())

                    # 格式校验
                    if json_data == [] or not isinstance(json_data[0], dict):
                        continue

                    for v in json_data:
                        # 格式校验
                        if "message" not in v:
                            continue

                        if v.get("message") != "":
                            items.append(
                                CacheItem({
                                    "src": v.get("message"),
                                    "dst": v.get("message"),
                                    "extra_field": v.get("name", ""),
                                    "row": len(items),
                                    "file_type": CacheItem.FileType.MESSAGEJSON,
                                    "file_path": rel_path,
                                })
                            )

        return items

    # Message JSON
    def write_to_path_messagejson(self, input_path: str, output_path: str, items: list[CacheItem]) -> None:
        # [
        #     {
        #         "message": "<fgName:pipo-fog004><fgLoopX:1><fgLoopY:1><fgSx:-2><fgSy:0.5>"
        #     },
        #     {
        #         "message": "エンディングを変更しますか？"
        #     },
        #     {
        #         "message": "はい"
        #     },
        # ]

        target = [
            item for item in items
            if item.get_file_type() == CacheItem.FileType.MESSAGEJSON
        ]

        data: dict[str, list[str]] = {}
        for item in target:
            data.setdefault(item.get_file_path(), []).append(item)

        for rel_path, items in data.items():
            abs_path = os.path.join(output_path, rel_path)
            os.makedirs(os.path.dirname(abs_path), exist_ok = True)

            result = []
            for item in items:
                result.append({
                    "name": item.get_extra_field(),
                    "message": item.get_dst(),
                })

            with open(abs_path, "w", encoding = "utf-8") as writer:
                writer.write(json.dumps(result, indent = 4, ensure_ascii = False))